import pandas as pd
import os

file_path = r"d:\__SELF\Feedback_management_system\Backend\scratch\data\RO_List.xlsx"

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
else:
    try:
        df = pd.read_excel(file_path)
        print("Columns:", df.columns.tolist())
        print("First 5 rows:")
        print(df.head().to_markdown(index=False, numalign="left", stralign="left"))
        print(f"\nTotal rows: {len(df)}")
    except ImportError:
        print("pandas or openpyxl not installed.")
        # Try openpyxl directly if pandas fails
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            sheet = wb.active
            print("Sheet names:", wb.sheetnames)
            print("First 5 rows (raw):")
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i < 5:
                    print(row)
                else:
                    break
        except ImportError:
            print("openpyxl also not installed.")
    except Exception as e:
        print(f"Error reading file: {e}")
